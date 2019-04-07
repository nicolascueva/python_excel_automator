
# Make sure to download openpyxl module
import openpyxl as xl
from openpyxl.chart import BarChart, Reference

# function to run with file passed in
def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    # cell = ['a1']
    # cell = sheet.cell(1, 1)

    # Loop through the rows to get prices then update them
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    # Select where your bar graph goes
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'E2')

    wb.save(filename)

#pass in file you want
process_workbook('transactions.xlsx')



